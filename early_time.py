"""
    ================================================================
    A python script to investigate the early time, non steady-state 
    behaviour of mRNA translation. This research was conducted as 
    part of my Senior Honours Project at the University of Edinburgh
    ================================================================
    Author: C. Abbott
    Version: Feb 2020
    ================================================================
"""
from ProteinSynthesis import ProteinSynthesis
import sys
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

def main():
    # Taking in arguments.
    if len(sys.argv) != 3:
        print("Wrong number of arguments.")
        print("Usage: " + sys.argv[0] +
              " <parameters file>" + "<translation rates>")
        quit()
    else:
        simul_parameters = sys.argv[1]
        trans_params = sys.argv[2]

    # Loading Excel spreadsheet.
    wb = openpyxl.load_workbook("Initiation-Rates.xlsx")
    # Choosing appropriate sheet.
    ws = wb['LacZ']
    # Storing genes and alphas.
    alphas = []
    genes = []
    # Get alphas.
    for column in ws['B2':'B786']:
        for cell in column:
            alphas.append(cell.value)
    # Get gene names.
    for column in ws['A2':'A786']:
        for cell in column:
            genes.append(cell.value)
    # Creating dictionary to hold alphas.
    alpha_dict = dict(zip(genes, alphas))

    # Open input file and assinging parameters.
    with open(str(simul_parameters), "r") as f:
        # Read the lines of the input data file.
        line = f.readline()
        items = line.split(", ")
        l = int(items[0])        # Ribosome length.
        n_traj = int(items[1])   # Number of trajectories.
        n_meas = int(items[2])   # Number of measurements.
    alpha = alpha_dict[str(trans_params[15:19])]

    # Initialising elongation rates.
    with open(str(trans_params), "r") as f:
        elong_omegas = np.array([float(line.split()[1]) for line in f])
    omegas = np.zeros(elong_omegas.size + 1)
    omegas[1:] = elong_omegas  # Adding elongation rates.
    omegas[0] = alpha          # Initiation rate.
    L = int(omegas.size)       # Setting length of mRNA.

    # Time taken for first ribosome to complete translation.
    exact_t1 = np.sum(np.reciprocal(omegas[1:]))
    # Upper time limit for simulation. 
    T = 1.5 * exact_t1
    
    # Setting time domains for observables.
    dt = T / n_meas
    measure_times = np.arange(0, T, dt)
    # Density data storage.
    densities = np.zeros(measure_times.size)
    # Tagged ribosome times to complete synthesis of first peptide.
    tagged_times = []

    # Simulations begin.
    for i in range(n_traj):
        print(i)
        # Density data storage.
        traj_densities = np.zeros(measure_times.size)
        # Tracking first ribosome.
        ribosome_pos = 1
        check = False
        # Time initiation.
        t_old = 0
        t_new = 0
        # Tick finding.
        k1 = 0
        k2 = 0
        # Create new instance of the simulation for every trajectory.
        simulation = ProteinSynthesis(
            length=l, size=L, alpha=alpha, omegas=omegas)
        # Begin trajectory.
        while t_new <= T:
            k1 = simulation.get_k1(t_old, dt) # Get k1.
            # Collect all possible moves.
            R = simulation.get_R()
            # Sample random time.
            t_new += simulation.get_random_time(R)
            k2 = simulation.get_k2(t_new, dt) # Get k2.
            # Final loop check.
            if t_new > T:
                k2 = n_meas
            # Setting values.
            traj_densities[k1:k2+1] = simulation.get_occupation_number()
            # Update t_old
            t_old = t_new
            # Choose which ribosome moves.
            index = simulation.get_transition(R)
            # Update simulation - ribosome hops.
            simulation.update(index)
            # Tracking initial ribosome position.
            if check == False:
                if (index+1) == ribosome_pos:
                    ribosome_pos += 1
                elif ribosome_pos > simulation.size:
                    tagged_times.append(t_new)
                    check = True
        # Collect data for each trajectory. 
        densities += traj_densities
    
    # Compute actual density.
    densities = densities / (n_traj * (L - 1))
    
    # Comparing analytical and experimental results
    avg_t1 = np.mean(np.array(tagged_times))
    print(avg_t1)
    print(exact_t1)
    
    # Plotting.
    simulation.plot_density(measure_times, densities,
                            avg_t1, str(trans_params[15:19]))
    # Saving data.
    simulation.save_data(measure_times, densities, str(trans_params[15:19]))

main()
        

        
        
    

    

